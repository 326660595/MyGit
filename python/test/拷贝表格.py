# 如果您要将数据从Excel表格（a表格）的第二到第五行复制到Excel表格（b表格）的第八行后面，您可
# 以在openpyxl库中使用insert_rows方法来实现。以下是一个示例代码，演示如何执行此操作：
import openpyxl

# 打开源Excel文件（a表格）和目标Excel文件（b表格）
source_workbook = openpyxl.load_workbook('a.xlsx')
target_workbook = openpyxl.load_workbook('b.xlsx')

# 选择要从源工作表复制的数据
source_sheet = source_workbook.active

# 选择要将数据粘贴到的目标工作表
target_sheet = target_workbook.active

# 获取源表格的第二到第五行数据
data_to_copy = []
for row_index in range(2, 6):  # 复制第二到第五行
    row_data = [cell.value for cell in source_sheet[row_index]]
    data_to_copy.append(row_data)

# 插入空行到目标表格的第八行后面
target_sheet.insert_rows(8, amount=len(data_to_copy))

# 将数据粘贴到插入的行
for row_data, target_row in zip(data_to_copy, target_sheet.iter_rows(min_row=8, max_row=8+len(data_to_copy), min_col=1, max_col=len(data_to_copy[0]))):
    for source_value, target_cell in zip(row_data, target_row):
        target_cell.value = source_value

# 保存目标Excel文件
target_workbook.save('b.xlsx')

# 关闭工作簿
source_workbook.close()
target_workbook.close()
