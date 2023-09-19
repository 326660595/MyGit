from openpyxl import Workbook

# 创建一个新的Excel文件
workbook = Workbook()

# 获取默认的工作表（sheet1）
sheet1 = workbook.create_sheet(title='Sheet1')

# 在sheet1的A1单元格写入"s1"
sheet1['A1'] = 's1'

# 新增一个工作表（sheet2）
sheet2 = workbook.create_sheet(title='Sheet2')

# 在sheet2的A1单元格写入"s2"
sheet2['A1'] = 's2'

# 保存Excel文件
excel_file = 'path_to_save_excel_file.xlsx'  # 替换为你想保存Excel文件的路径
workbook.save(filename=excel_file)

# 关闭Excel文件
workbook.close()
