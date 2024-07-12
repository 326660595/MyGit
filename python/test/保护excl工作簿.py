from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.protection import SheetProtection

# 加载 Excel 文件
workbook = load_workbook('C:\\code\\MyGit\\python\\test\\config_log_check.xlsx')

# 获取需要保护的工作表
worksheet = workbook.active
# 或者通过指定工作表名称加载
# worksheet = workbook['Sheet1']

# 创建 SheetProtection 对象
protection = SheetProtection(sheet=True, password='123')

# 应用保护设置到工作表
worksheet.protection = protection

# 保存 Excel 文件
workbook.save('protected_file.xlsx')
# 1
# l7lz3nq5
